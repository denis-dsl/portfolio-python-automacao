"""Normalização e formatação de planilhas Excel."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.run_artifacts import RunSummary, make_run_dir, write_summary
from src.schema import COLUMN_ALIASES, REQUIRED_COLUMNS, canon_col
from src.validation import has_errors, validate_dataframe, write_issues_csv

_RE_DATE = re.compile(r"(\d{2}/\d{2}/\d{4})")


def _normalize_yes_no(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip().upper()
    if s in {"SIM", "S", "YES", "Y", "TRUE", "1"}:
        return "SIM"
    if s in {"NÃO", "NAO", "N", "NO", "FALSE", "0", ""}:
        return "NÃO" if s else ""
    return s


def _parse_date_mixed(value: Any) -> pd.Timestamp | None:
    """Converte datas em formatos mistos para Timestamp (ou None).
    Aceita:
    - '02/01/2026'
    - '() 02/01/2026'
    - '2026-01-02'
    - '2026-01-02T00:00:00'
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    m = _RE_DATE.search(s)
    if m:
        ts = pd.to_datetime(m.group(1), dayfirst=True, errors="coerce")
        return None if pd.isna(ts) else ts

    ts = pd.to_datetime(s, errors="coerce")
    return None if pd.isna(ts) else ts


def _parse_money_mixed(value: Any) -> float | None:
    """
    Converte strings tipo:
    - '3.000,00' (pt-BR)
    - '3,000.00' (en-US)
    - '3000.00'
    - '3000'
    Para float.
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    # remove R$, espaços etc.
    s = s.replace("R$", "").replace(" ", "")

    # Se tem . e ,:
    # - pt: 1.234,56 => remove '.' e troca ',' por '.'
    # - en: 1,234.56 => remove ',' e mantém '.'
    if "." in s and "," in s:
        last_dot = s.rfind(".")
        last_comma = s.rfind(",")
        if last_comma > last_dot:  # pt
            s = s.replace(".", "").replace(",", ".")
        else:  # en
            s = s.replace(",", "")

    # Se só tem vírgula, assume decimal pt: 123,45
    elif "," in s and "." not in s:
        s = s.replace(".", "").replace(",", ".")

    # Se só tem ponto ou nada, ok
    try:
        return float(s)
    except ValueError:
        return None


def normalize_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Normaliza colunas (datas, valores e flags) de um DataFrame."""
    df = df.copy()

    # Padronizar nomes de colunas (tira espaços duplicados)
    df.columns = [str(c).strip() for c in df.columns]

    # Renomear colunas para o padrão interno usando aliases (mundo real)
    rename_map = {}
    for col in df.columns:
        key = canon_col(col)
        if key in COLUMN_ALIASES:
            rename_map[col] = COLUMN_ALIASES[key]

    df = df.rename(columns=rename_map)

    missing = sorted(REQUIRED_COLUMNS - set(df.columns))
    if missing:
        # deixa o DataFrame seguir? aqui a gente prefere quebrar já
        raise ValueError(f"Colunas obrigatórias ausentes: {missing}")

    # Padronizar nomes finais (tira espaços duplicados)
    df.columns = [str(c).strip() for c in df.columns]

    stats = {}

    # Normalizações específicas (se existirem)
    if "Taxa" in df.columns:
        df["Taxa"], stats["Taxa"] = _apply_with_stats(df["Taxa"], _parse_money_mixed)

    if "Emissão" in df.columns:
        df["Emissão"], stats["Emissão"] = _apply_with_stats(df["Emissão"], _parse_date_mixed)

    if "Data Pagamento" in df.columns:
        df["Data Pagamento"] = df["Data Pagamento"].apply(_parse_date_mixed)

    for col in ["Vlr Liberado", "Valor Cancelado", "VALOR PAGO", "Valor Devolvido"]:
        if col in df.columns:
            df[col] = df[col].apply(_parse_money_mixed)

    if "Pago" in df.columns:
        df["Pago"] = df["Pago"].apply(_normalize_yes_no)

    if "TED/Devolvida" in df.columns:
        df["TED/Devolvida"] = df["TED/Devolvida"].apply(_normalize_yes_no)

    # Limpar contrato (garantir string)
    if "Contrato" in df.columns:
        df["Contrato"] = df["Contrato"].astype(str).str.strip()

    # Substituir NaN por None
    df = df.where(pd.notnull(df), None)
    return df, stats


def _format_excel(path: Path) -> None:
    """
    Formatação openpyxl:
    - Congela primeira linha
    - Centraliza cabeçalho
    - Ajusta largura de colunas
    - Formata datas e moedas
    """
    wb = load_workbook(path)
    ws = wb.active

    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    # alinhamento padrão
    body_alignment = Alignment(vertical="top")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment

    # formatos por coluna (baseado no nome do header)
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    date_cols = {"Emissão", "Data Pagamento"}
    money_cols = {"Vlr Liberado", "Valor Cancelado", "VALOR PAGO", "Valor Devolvido"}
    rate_cols = {"Taxa"}
    int_cols = {"Prazo"}
    text_cols = {"Contrato"}

    for h, idx in headers.items():
        col_letter = get_column_letter(idx)
        if h in date_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = "DD/MM/YYYY"
        if h in money_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = "#,##0.00"
        if h in rate_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = "0.00"
        if h in int_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = "0"
        if h in text_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = "@"

    # auto width simples
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(path)


def normalize_excel(input_path: Path, output_path: Path, strict: bool = False) -> None:
    """Lê um Excel, normaliza os dados e salva artifacts de execução."""

    logger.info("Carregando Excel com pandas...")
    df = pd.read_excel(input_path)

    logger.info(f"Linhas lidas: {len(df)}")
    df2, transform_stats = normalize_dataframe(df)

    # Cria pasta da execução
    run_dir = make_run_dir(output_path.parent)
    run_output = run_dir / output_path.name
    issues_path = run_dir / "issues.csv"
    summary_path = run_dir / "summary.json"

    logger.info("Validando inconsistências (WARN/ERROR)...")
    issues = validate_dataframe(df2)
    has_any = write_issues_csv(issues, issues_path)

    num_errors = sum(1 for i in issues if i.severity == "ERROR")
    num_warn = sum(1 for i in issues if i.severity == "WARN")
    logger.info(f"Métricas: {num_errors} ERROR, {num_warn} WARN")

    if has_any:
        logger.warning(f"Issues encontradas. Verifique: {issues_path.resolve()}")
        if strict and has_errors(issues):
            raise ValueError("Execução em modo --strict: issues ERROR detectadas.")

    logger.info("Salvando Excel normalizado...")
    df2.to_excel(run_output, index=False)

    logger.info("Aplicando formatação openpyxl...")
    _format_excel(run_output)

    summary = RunSummary(
        started_at=datetime.now().isoformat(timespec="seconds"),
        input_file=str(input_path),
        output_file=str(run_output),
        rows=int(len(df2)),
        cols=int(len(df2.columns)),
        num_errors=num_errors,
        num_warn=num_warn,
        notes={
            "transform_stats": transform_stats,
        },
    )

    write_summary(summary, summary_path)

    logger.info(f"Artifacts salvos em: {run_dir.resolve()}")
    logger.info("Concluído.")


def generate_validation_report(df: pd.DataFrame, path: Path) -> bool:
    """Gera CSV com inconsistências básicas."""
    errors = []

    for idx, row in df.iterrows():
        if row.get("Pago") == "SIM":
            if pd.isna(row.get("Data Pagamento")):
                errors.append((idx, "Pago = SIM mas Data Pagamento vazio"))
            if pd.isna(row.get("VALOR PAGO")):
                errors.append((idx, "Pago = SIM mas Valor Pago vazio"))

        contrato = row.get("Contrato")
        if contrato is None or str(contrato).strip() == "":
            errors.append((idx, "Contrato vazio"))

        if row.get("TED/Devolvida") == "SIM" and pd.isna(row.get("Valor Devolvido")):
            errors.append((idx, "TED/Devolvida = SIM mas Valor Devolvido vazio"))

        if row.get("Pago") != "SIM" and (not pd.isna(row.get("VALOR PAGO"))):
            errors.append((idx, "Pago != SIM mas VALOR PAGO preenchido"))

    if errors:
        err_df = pd.DataFrame(errors, columns=["Linha", "Erro"])
        path.parent.mkdir(parents=True, exist_ok=True)
        err_df.to_csv(path, index=False)
        return True

    if path.exists():
        path.unlink()

    return False


def _apply_with_stats(series: pd.Series, func) -> tuple[pd.Series, dict[str, int]]:
    """Aplica func e conta sucesso/falha (None)."""
    before_non_empty = series.notna().sum()
    out = series.apply(func)
    after_non_empty = pd.Series(out).notna().sum()
    return out, {
        "before_non_empty": int(before_non_empty),
        "after_non_empty": int(after_non_empty),
        "became_empty": int(before_non_empty - after_non_empty),
    }
